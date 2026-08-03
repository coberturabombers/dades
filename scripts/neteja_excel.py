#!/usr/bin/env python3
"""
neteja_excel.py - Genera una copia NETA de l'Excel del TRI
Nomes conserva parc, minim de torn i reals de les 7 regions.
Us: python neteja_excel.py "Personal de guardia.xlsx"
Genera: cobertura_neta.xlsx
"""
import sys
import os

try:
    import openpyxl
except ImportError:
    print("ERROR: cal instal-lar openpyxl  ->  pip install openpyxl", file=sys.stderr)
    sys.exit(1)

REGION_START_COLS = {2: "REC", 14: "REG", 26: "REMN", 38: "REL",
                     50: "REMS", 62: "RET", 74: "RETE"}


def main():
    if len(sys.argv) >= 2:
        src = sys.argv[1]
    else:
        src = input("Arrossega aqui l'Excel complet i prem Enter:\n> ").strip().strip('"').strip("'")
    if not os.path.exists(src):
        print("No trobo el fitxer: " + src)
        input("Prem Enter per sortir.")
        return
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb["Dades"]
    out = openpyxl.Workbook()
    ods = out.active
    ods.title = "Dades"
    for start_col, region in REGION_START_COLS.items():
        ods.cell(row=2, column=start_col, value=region)
        ods.cell(row=3, column=start_col, value="PARC")
        ods.cell(row=3, column=start_col + 1, value="MINIM TORN")
        ods.cell(row=3, column=start_col + 2, value="REALS A PARC")
        for r in range(4, 60):
            parc = ws.cell(row=r, column=start_col).value
            mn = ws.cell(row=r, column=start_col + 1).value
            rl = ws.cell(row=r, column=start_col + 2).value
            if parc is not None:
                ods.cell(row=r, column=start_col, value=parc)
            if mn is not None:
                ods.cell(row=r, column=start_col + 1, value=mn)
            if rl is not None:
                ods.cell(row=r, column=start_col + 2, value=rl)
    try:
        src_resum = wb["Resum"]
        ores = out.create_sheet("Resum")
        ores.cell(row=2, column=2, value="DATA")
        ores.cell(row=4, column=2, value=src_resum.cell(row=4, column=2).value)
    except Exception:
        pass
    dest = os.path.join(os.path.dirname(src) or ".", "cobertura_neta.xlsx")
    out.save(dest)
    print("\nFet! Copia neta desada a:\n  " + dest)
    input("\nPrem Enter per sortir.")


if __name__ == "__main__":
    main()
