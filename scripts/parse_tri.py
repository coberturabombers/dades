#!/usr/bin/env python3
"""
parse_tri.py - Motor de lectura del TRI (Assemblea Bombers de Catalunya)
========================================================================
VERSIÓ 2 - Llegeix el document mestre "Personal de guàrdia" en Google Sheets
(full VISOR), que conté les 7 regions completes, sense #REF!.

Fa servir la columna "A PARC" (TOTAL - GOLF): els efectius realment disponibles
al parc, descomptant els destinats a unitats Golf. És la mètrica de referència.

Ús:
    python parse_tri.py --url "<enllaç CSV del VISOR>"
    python parse_tri.py --input fitxer.csv
    python parse_tri.py --input fitxer.xlsx   (llegeix el full VISOR)
"""

import argparse
import csv
import io
import json
import os
import re
import sys
from datetime import date

# --- Estructura del full VISOR del document mestre -------------------------
# Columnes reals de l'Excel (openpyxl 1-based): PARC a 2,15,28,41,54,67,80.
# Aquí en índex 0-based (per treballar amb CSV i xlsx igual): PARC-1.
# Per cada regió: PARC (0-based), MÍNIM = +2, REALS A PARC = +3, A PARC = +4
REGION_COLS = {
    "REC": 1, "REG": 14, "REMN": 27, "REL": 40, "REMS": 53, "RET": 66, "RETE": 79,
}
REGION_ORDER = ["REC", "REG", "REMN", "REL", "REMS", "RET", "RETE"]
REGION_NAMES = {
    "REC": "Centre", "REG": "Girona", "REMN": "Metropolitana Nord",
    "REL": "Lleida", "REMS": "Metropolitana Sud", "RET": "Tarragona",
    "RETE": "Terres de l'Ebre",
}

# --- Selecció d'efectius --------------------------------------------------
# Prioritat 1: REALS A PARC del document mestre (+3).
# Prioritat 2: si REC/REG/REMN/REL surten buides, es busca REALS A PARC (columna H)
#   al seu Google Sheet regional PÚBLIC.
# Prioritat 3 (reserva): A PARC / TOTAL-GOLF (+4) del document mestre.
COL_OFFSET_REALS = 3
COL_OFFSET_TOTALGOLF = 4
COL_OFFSET_MINIM = 2

# Google Sheets regionals PÚBLICS (només aquests 4 són accessibles sense permís).
# La columna H conté REALS A PARC. gid del full de dades = 255080715.
# S'usen com a reserva quan el document mestre deixa la regió buida.
PUBLIC_REGIONAL_SHEETS = {
    "REC":  ("1q6ACcBeUgskxKs6WQFQQOU4iU8Sg0_b4_q3-MEnVbYk", 255080715),
    "REG":  ("1YDdfSR2bj5pQBHpVhOOYnzVx2D0ysHME7RJEnundGl4", 255080715),
    "REMN": ("1s2VuvKTbFMvtno4o1Wq3r95NTb49bK0Vwp1sHXz1574", 255080715),
    "REL":  ("11TNgNeQTuJLkE4kD0LBqbBe55ZOl_6DH2W95N56GgZ0", 255080715),
}

KNOWN_NOCODE = {"GROS"}
MIN_CHECKSUM = {"REC": 45, "REG": 71, "REMN": 69, "REL": 45,
                "REMS": 55, "RET": 42, "RETE": 27}

PARC_COORDS = {
    "03 BER": (42.104, 1.845), "08 GUA": (42.203, 1.857), "15 PUIG": (42.431, 1.928),
    "19 SOLS": (41.995, 1.517), "04 CAF": (41.734, 1.512), "05 CAR": (41.913, 1.681),
    "10 MAN": (41.728, 1.827), "11 MOI": (41.812, 2.096), "14 PRA": (42.010, 2.032),
    "20 TOR": (42.049, 2.259), "21 VIC": (41.930, 2.254),
    "10 FIGU": (42.267, 2.961), "14 LLAN": (42.363, 3.152), "22 ROSE": (42.262, 3.176),
    "26 TORR": (42.043, 3.127), "11 GIRO": (41.983, 2.824), "17 OLOT": (42.181, 2.489),
    "21 RIPO": (42.201, 2.190), "03 BANY": (42.119, 2.767), "15 LLOR": (41.700, 2.845),
    "16 MACA": (41.777, 2.734), "16 MAÇA": (41.777, 2.734), "25 SCFA": (41.860, 2.669), "01 AMER": (41.968, 2.601),
    "18 PALA": (41.918, 3.163), "28 VALL": (41.818, 3.033), "04 PERA": (41.981, 2.965),
    "08 CASS": (41.885, 2.874),
    "GROS": (41.386, 2.170), "13 RUB": (41.493, 2.033), "14 SAB": (41.548, 2.107),
    "20 TER": (41.563, 2.010), "06 GRA": (41.608, 2.288), "10 MOL": (41.540, 2.213),
    "16 SCE": (41.689, 2.491), "02 BAD": (41.450, 2.247), "18 SCG": (41.452, 2.208),
    "09 MAT": (41.538, 2.445), "12 PIN": (41.627, 2.689),
    "09 CERV": (41.671, 1.272), "19 LLEI": (41.615, 0.626), "20 MOLL": (41.632, 0.895),
    "31 TAR": (41.647, 1.140), "06 BALA": (41.790, 0.807), "29 SEU": (42.358, 1.461),
    "25 PONT": (42.408, 0.741), "30 SORT": (42.412, 1.128), "33 TREM": (42.166, 0.895),
    "04 COR": (41.354, 2.070), "05 GAV": (41.306, 2.000), "07 HOS": (41.360, 2.100),
    "12 PLL": (41.327, 2.095), "13 SBOI": (41.345, 2.037), "15 SFE": (41.383, 2.044),
    "08 IGU": (41.579, 1.617), "10 MAR": (41.474, 1.930), "18 VIF": (41.346, 1.699),
    "19 VIL": (41.224, 1.725),
    "05 MONT": (41.375, 1.161), "13 VALS": (41.287, 1.249), "12 TARR": (41.119, 1.245),
    "14 VEND": (41.220, 1.535), "02 CAMB": (41.067, 1.058), "04 FALS": (41.145, 0.821),
    "05 HOSP": (40.995, 0.938), "08 REUS": (41.156, 1.107),
    "22 ASCO": (41.183, 0.564), "26 GAND": (41.053, 0.437), "28 MORA": (41.093, 0.643), "28 MÓRA": (41.093, 0.643),
    "20 AMET": (40.884, 0.802), "21 AMPO": (40.708, 0.581), "30 TORT": (40.812, 0.521),
    "31 ULLD": (40.596, 0.451),
}

PARC_NAMES = {
    "03 BER": "Berga", "08 GUA": "Guardiola de B.", "15 PUIG": "Puigcerdà",
    "19 SOLS": "Solsona", "04 CAF": "Calaf", "05 CAR": "Cardona", "10 MAN": "Manresa",
    "11 MOI": "Moia", "14 PRA": "Prats de Llucanes", "20 TOR": "Torello", "21 VIC": "Vic",
    "10 FIGU": "Figueres", "14 LLAN": "Llanca", "22 ROSE": "Roses",
    "26 TORR": "Torroella de Montgri", "11 GIRO": "Girona", "17 OLOT": "Olot",
    "21 RIPO": "Ripoll", "03 BANY": "Banyoles", "15 LLOR": "Lloret de Mar",
    "16 MACA": "Maçanet de la Selva", "16 MAÇA": "Maçanet de la Selva", "25 SCFA": "Santa Coloma de Farners", "01 AMER": "Amer",
    "18 PALA": "Palafrugell", "28 VALL": "Vall d'Aro", "04 PERA": "La Pera",
    "08 CASS": "Cassa de la Selva",
    "GROS": "Barcelona (Gros)", "13 RUB": "Rubi", "14 SAB": "Sabadell",
    "20 TER": "Terrassa", "06 GRA": "Granollers", "10 MOL": "Mollet",
    "16 SCE": "Sant Celoni", "02 BAD": "Badalona", "18 SCG": "Santa Coloma de Gramanet",
    "09 MAT": "Mataro", "12 PIN": "Pineda",
    "09 CERV": "Cervera", "19 LLEI": "Lleida", "20 MOLL": "Mollerussa", "31 TAR": "Tarrega",
    "06 BALA": "Balaguer", "29 SEU": "La Seu d'Urgell", "25 PONT": "El Pont de Suert",
    "30 SORT": "Sort", "33 TREM": "Tremp",
    "04 COR": "Cornella de Ll.", "05 GAV": "Gava", "07 HOS": "L'Hospitalet de Ll.",
    "12 PLL": "El Prat de Llobregat", "13 SBOI": "Sant Boi de Ll.", "15 SFE": "S. Feliu de Ll.",
    "08 IGU": "Igualada", "10 MAR": "Martorell", "18 VIF": "Vilafranca del P.",
    "19 VIL": "Vilanova i la Geltru",
    "05 MONT": "Montblanc", "13 VALS": "Valls", "12 TARR": "Tarragona",
    "14 VEND": "El Vendrell", "02 CAMB": "Cambrils", "04 FALS": "Falset",
    "05 HOSP": "L'Hospitalet de l'Infant", "08 REUS": "Reus",
    "22 ASCO": "Asco", "26 GAND": "Gandesa", "28 MORA": "Móra d'Ebre", "28 MÓRA": "Móra d'Ebre",
    "20 AMET": "L'Ametlla de Mar", "21 AMPO": "Amposta", "30 TORT": "Tortosa",
    "31 ULLD": "Ulldecona",
}


def clean_code(value):
    """Neteja el codi de parc: treu emojis i espais extra."""
    if not value:
        return None
    s = str(value).strip()
    # treu emojis i símbols (🔴🔵⏬⏫❌ etc.)
    s = re.sub(r'[^\w\sÀ-ÿ]', '', s).strip()
    return s if s else None


def is_parc(code):
    if not code:
        return False
    if code in KNOWN_NOCODE:
        return True
    return bool(re.match(r"^\d{2}\s+[A-Z]", code))


def to_int(v):
    try:
        return int(float(str(v).strip()))
    except (ValueError, TypeError):
        return None


def read_rows_from_csv(text):
    return list(csv.reader(io.StringIO(text)))


def read_rows_from_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["VISOR"] if "VISOR" in wb.sheetnames else wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c) for c in row])
    return rows


def extract_date(rows):
    """Busca la data de previsió al text del document."""
    for row in rows[:2]:
        line = " ".join(str(c) for c in row)
        m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", line)
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return f"{y:04d}-{mo:02d}-{d:02d}"
    return None


def fetch_public_reals(region):
    """Baixa la columna H (REALS A PARC) del Google Sheet regional públic.
    Retorna una llista de valors (int o None) en ordre, o None si falla."""
    if region not in PUBLIC_REGIONAL_SHEETS:
        return None
    sheet_id, gid = PUBLIC_REGIONAL_SHEETS[region]
    url = ("https://docs.google.com/spreadsheets/d/" + sheet_id +
           "/export?format=csv&gid=" + str(gid))
    try:
        import urllib.request
        opener = urllib.request.build_opener()
        opener.addheaders = [("User-Agent", "Mozilla/5.0")]
        data = opener.open(url, timeout=30).read().decode("utf-8", "replace")
        rows = list(csv.reader(io.StringIO(data)))
        # Columna H = índex 7 (0-based). Retornem tots els valors numèrics d'aquesta columna.
        vals = []
        for r in rows:
            if len(r) > 7:
                vals.append(to_int(r[7]))
            else:
                vals.append(None)
        return vals
    except Exception:
        return None


def parse_rows(rows):
    parcs = []
    warnings = []
    n_reals = 0
    n_public = 0
    n_fallback = 0
    for region, base in REGION_COLS.items():
        col_min = base + COL_OFFSET_MINIM
        col_reals = base + COL_OFFSET_REALS
        col_golf = base + COL_OFFSET_TOTALGOLF
        region_min_sum = 0
        # Primer, recollim els parcs de la regió del document mestre
        region_parcs = []
        need_public = False
        for row in rows:
            if base >= len(row):
                continue
            code = clean_code(row[base])
            if not is_parc(code):
                continue
            mn = to_int(row[col_min]) if col_min < len(row) else None
            reals = to_int(row[col_reals]) if col_reals < len(row) else None
            golf = to_int(row[col_golf]) if col_golf < len(row) else None
            if reals is None:
                need_public = True
            region_parcs.append({"code": code, "min": mn, "reals": reals, "golf": golf})
            if mn is not None:
                region_min_sum += mn
        # Si falten REALS i la regió té Sheet públic, els busquem allà
        public_vals = None
        if need_public and region in PUBLIC_REGIONAL_SHEETS:
            public_vals = fetch_public_reals(region)
        # Assignem l'efectiu de cada parc segons prioritat
        for idx, rp in enumerate(region_parcs):
            code, mn, reals, golf = rp["code"], rp["min"], rp["reals"], rp["golf"]
            if reals is not None:
                ef = reals
                n_reals += 1
            elif public_vals is not None and idx < len(public_vals) and public_vals[idx] is not None:
                ef = public_vals[idx]
                n_public += 1
            else:
                ef = golf
                if golf is not None:
                    n_fallback += 1
            lat, lon = PARC_COORDS.get(code, (None, None))
            # Mobilitat: entren = TOTAL-GOLF (dotació a les 7am), queden = REALS A PARC
            entren = golf
            queden = reals if reals is not None else (public_vals[idx] if (public_vals is not None and idx < len(public_vals)) else None)
            mobilitat = None
            if entren is not None and queden is not None:
                mobilitat = queden - entren  # negatiu = han marxat, positiu = han rebut
            parcs.append({
                "code": code, "name": PARC_NAMES.get(code, code),
                "region": region, "min": mn, "real": ef, "lat": lat, "lon": lon,
                "entren": entren, "queden": queden, "mobilitat": mobilitat,
            })
        expected = MIN_CHECKSUM.get(region)
        if expected is not None and region_min_sum != expected:
            warnings.append(
                "ATENCIO: suma de minims de " + region + " = " + str(region_min_sum) +
                ", s'esperava " + str(expected) + ".")
    info = str(n_reals) + " REALS (mestre)"
    if n_public:
        info += ", " + str(n_public) + " REALS (Sheet públic)"
    if n_fallback:
        info += ", " + str(n_fallback) + " TOTAL-GOLF (reserva)"
    warnings.append("INFO: " + info)
    return parcs, warnings


def build_snapshot(the_date, parcs):
    regions = {rc: {"sobre_minims": 0, "minims": 0, "inacceptable": 0,
                    "greu": 0, "critic": 0, "tancat": 0} for rc in REGION_ORDER}
    total_min = total_real = sota_minims = tancats = 0
    dotacions_falten = 0
    for p in parcs:
        rc = p["region"]
        mn, rl = p["min"], p["real"]
        if mn is None or rl is None:
            continue
        total_min += mn
        total_real += rl
        falten = mn - rl
        if falten > 0:
            dotacions_falten += falten
        if rl == 0:
            regions[rc]["tancat"] += 1
            tancats += 1
            sota_minims += 1
        elif falten <= 0:
            if rl > mn:
                regions[rc]["sobre_minims"] += 1
            else:
                regions[rc]["minims"] += 1
        else:
            sota_minims += 1
            if falten == 1:
                regions[rc]["inacceptable"] += 1
            elif falten == 2:
                regions[rc]["greu"] += 1
            else:
                regions[rc]["critic"] += 1
    # Estat INICIAL (a les 7h, amb els que entren = TOTAL-GOLF) vs FINAL (amb REALS)
    sota_inici = 0
    sobre_inici = 0
    desplacats = 0
    for p in parcs:
        entren = p.get("entren")
        mob = p.get("mobilitat")
        mn = p["min"]
        if entren is not None and mn is not None:
            if entren < mn:
                sota_inici += 1
            elif entren > mn:
                sobre_inici += 1
        if mob is not None and mob < 0:
            desplacats += -mob
    return {
        "date": the_date, "parcs": parcs, "regions": regions,
        "catalunya": {
            "min_total": total_min, "real_total": total_real,
            "diferencia": total_real - total_min,
            "dotacions_falten": dotacions_falten,
            "sota_minims": sota_minims,
            "tancats": tancats,
            "sota_inici": sota_inici,
            "sobre_inici": sobre_inici,
            "desplacats": desplacats,
            "n_parcs": len([p for p in parcs if p["real"] is not None]),
        },
    }


def download(url, dest):
    import urllib.request
    import urllib.parse
    import http.cookiejar
    cj = http.cookiejar.CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
    opener.addheaders = [("User-Agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64)")]
    resp = opener.open(url)
    data = resp.read()
    # Si el que rebem és HTML (pàgina de confirmació/login de Google), intentem
    # seguir el formulari de confirmació de descàrrega.
    head = data[:512].lstrip()[:15].lower()
    if head.startswith(b"<!doctype html") or head.startswith(b"<html"):
        text = data.decode("utf-8", "replace")
        m_action = re.search(r'action="([^"]+)"', text)
        confirm = re.search(r'name="confirm"\s+value="([^"]+)"', text)
        uuid = re.search(r'name="uuid"\s+value="([^"]+)"', text)
        if m_action and confirm:
            params = {"confirm": confirm.group(1)}
            if uuid:
                params["uuid"] = uuid.group(1)
            action = m_action.group(1).replace("&amp;", "&")
            new_url = action + ("&" if "?" in action else "?") + urllib.parse.urlencode(params)
            data = opener.open(new_url).read()
    with open(dest, "wb") as f:
        f.write(data)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input")
    ap.add_argument("--url")
    ap.add_argument("--date")
    ap.add_argument("--history", default="data/history.json")
    args = ap.parse_args()

    path = args.input
    if args.url:
        path = "/tmp/_tri_download"
        print("Descarregant document des de l'enllac...")
        download(args.url, path)

    if not path or not os.path.exists(path):
        print("ERROR: cal --input o --url valid.", file=sys.stderr)
        sys.exit(1)

    # detecta si és CSV o XLSX
    with open(path, "rb") as f:
        head = f.read(4)
    if head[:2] == b"PK":  # xlsx (zip)
        rows = read_rows_from_xlsx(path)
    else:
        with open(path, encoding="utf-8", errors="replace") as f:
            rows = read_rows_from_csv(f.read())

    parcs, warnings = parse_rows(rows)
    for w in warnings:
        print(w, file=sys.stderr)

    the_date = args.date or extract_date(rows) or date.today().isoformat()
    print("Data de la foto: " + the_date + "  (" + str(len(parcs)) + " parcs llegits)")

    snapshot = build_snapshot(the_date, parcs)
    history = {}
    if os.path.exists(args.history):
        with open(args.history, encoding="utf-8") as f:
            history = json.load(f)
    history[the_date] = snapshot
    os.makedirs(os.path.dirname(args.history) or ".", exist_ok=True)
    with open(args.history, "w", encoding="utf-8") as f:
        json.dump(history, f, ensure_ascii=False, separators=(",", ":"))
    latest_path = os.path.join(os.path.dirname(args.history) or ".", "latest.json")
    with open(latest_path, "w", encoding="utf-8") as f:
        json.dump(snapshot, f, ensure_ascii=False, separators=(",", ":"))

    cat = snapshot["catalunya"]
    print("OK. Sota minims: " + str(cat['sota_minims']) + "  Tancats: " +
          str(cat['tancats']) + "  Diferencia: " + str(cat['diferencia']) +
          "  (" + str(cat['n_parcs']) + " parcs)")


if __name__ == "__main__":
    main()
